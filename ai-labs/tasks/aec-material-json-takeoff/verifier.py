"""Verify the AEC material takeoff: a result JSON of cost totals plus a shareable .xlsx summary.

The ground truth is RECOMPUTED here from inputs/materials.json (staged to the agent) by the same
deterministic rules stated in the prompt, then cross-checked against expected/takeoff.json (never
staged). The submitted totals must match the recompute within MONEY_TOL; the missing-spec list must
match the recomputed set exactly. The artifact must be a real openpyxl workbook (not a renamed CSV)
that surfaces the trade names and the grand total as cells -- a structural key-value check, not a
re-derivation of the layout.

DETERMINISTIC RULES (single source of truth, mirrored verbatim in the prompt):
  - Exclude any line with status == "void" from ALL totals and from the missing-spec list.
  - line_cost = parse_quantity(quantity) * unit_cost, where parse_quantity strips commas and parses
    to a float.
  - totals_by_trade: sum of line_cost grouped by exact `trade` string.
  - totals_by_level: sum of line_cost grouped by exact `level` string.
  - grand_total_cost: sum of all non-void line_cost.
  - missing_spec_material_ids: the material_id of every non-void line whose spec_section is null or
    empty/whitespace. (Missing-spec lines STILL count toward cost totals -- you still owe the money.)
"""

import io

import openpyxl

from bench_verifier import fixtures, output, read_fixture_json, result

MONEY_TOL = 0.01


def parse_quantity(quantity: str) -> float:
    return float(str(quantity).replace(",", ""))


def _materials() -> list[dict]:
    # inputs/ is staged to the agent; recompute the oracle from it rather than trusting expected/.
    return read_fixture_json("inputs", "materials.json")


def _recompute(materials: list[dict]) -> dict:
    totals_by_trade: dict[str, float] = {}
    totals_by_level: dict[str, float] = {}
    missing_spec_material_ids: list[str] = []
    grand_total_cost = 0.0

    for item in materials:
        if item.get("status") == "void":
            continue
        line_cost = parse_quantity(item["quantity"]) * item["unit_cost"]
        totals_by_trade[item["trade"]] = totals_by_trade.get(item["trade"], 0.0) + line_cost
        totals_by_level[item["level"]] = totals_by_level.get(item["level"], 0.0) + line_cost
        grand_total_cost += line_cost
        spec = item["spec_section"]
        if spec is None or not str(spec).strip():
            missing_spec_material_ids.append(item["material_id"])

    return {
        "totals_by_trade": totals_by_trade,
        "totals_by_level": totals_by_level,
        "missing_spec_material_ids": set(missing_spec_material_ids),
        "grand_total_cost": grand_total_cost,
    }


def _expected() -> dict:
    rec = _recompute(_materials())
    # Cross-check the recompute against the committed verifier-only ground truth.
    if fixtures("expected", "takeoff.json").exists():
        stored = read_fixture_json("expected", "takeoff.json")
        assert abs(stored["grand_total_cost"] - rec["grand_total_cost"]) <= MONEY_TOL, (
            f"expected/takeoff.json grand_total {stored['grand_total_cost']} disagrees with recompute "
            f"{rec['grand_total_cost']:.2f}; fixtures out of sync"
        )
        assert set(stored["missing_spec_material_ids"]) == rec["missing_spec_material_ids"], (
            f"expected/takeoff.json missing-spec set {sorted(stored['missing_spec_material_ids'])} "
            f"disagrees with recompute {sorted(rec['missing_spec_material_ids'])}; fixtures out of sync"
        )
        for label, rec_totals in (("trade", rec["totals_by_trade"]), ("level", rec["totals_by_level"])):
            stored_totals = stored[f"totals_by_{label}"]
            assert set(stored_totals) == set(rec_totals), (
                f"expected/takeoff.json {label} set disagrees with recompute; fixtures out of sync"
            )
            for key, want in rec_totals.items():
                assert abs(stored_totals[key] - want) <= MONEY_TOL, (
                    f"expected/takeoff.json {label}={key!r} disagrees with recompute; fixtures out of sync"
                )
    return rec


def _submitted_map(rows: list, key: str) -> dict[str, float]:
    assert isinstance(rows, list), f"{key} block must be a list, got {type(rows).__name__}"
    out: dict[str, float] = {}
    for row in rows:
        name = row[key]
        assert name not in out, f"duplicate {key} in submission: {name!r}"
        out[name] = row["total_cost"]
    return out


def _assert_totals(label: str, key: str, submitted_rows: list, expected: dict[str, float]) -> None:
    got = _submitted_map(submitted_rows, key)
    assert set(got) == set(expected), (
        f"{label}: submitted {key}s {sorted(got)} != expected {sorted(expected)}"
    )
    for name, want in expected.items():
        assert abs(got[name] - want) <= MONEY_TOL, (
            f"{label}: {key}={name!r} total {got[name]} != expected {want:.2f}"
        )


# === result JSON checks ===

def test_totals_by_trade() -> None:
    expected = _expected()
    _assert_totals("totals_by_trade", "trade", result()["totals_by_trade"], expected["totals_by_trade"])


def test_totals_by_level() -> None:
    expected = _expected()
    _assert_totals("totals_by_level", "level", result()["totals_by_level"], expected["totals_by_level"])


def test_missing_spec_material_ids() -> None:
    expected = _expected()["missing_spec_material_ids"]
    submitted = result()["missing_spec_material_ids"]
    assert isinstance(submitted, list), f"missing_spec_material_ids must be a list, got {type(submitted).__name__}"
    assert set(submitted) == expected, (
        f"missing_spec_material_ids {sorted(set(submitted))} != expected {sorted(expected)}"
    )


def test_grand_total_cost() -> None:
    want = _expected()["grand_total_cost"]
    got = result()["grand_total_cost"]
    assert abs(got - want) <= MONEY_TOL, f"grand_total_cost {got} != expected {want:.2f}"


# === xlsx artifact checks ===

def _numeric_cell_values(ws) -> list[float]:
    values: list[float] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, bool):
                continue
            if isinstance(value, (int, float)):
                values.append(float(value))
            elif isinstance(value, str):
                stripped = value.replace("$", "").replace(",", "").strip()
                try:
                    values.append(float(stripped))
                except ValueError:
                    continue
    return values


def _text_cells(ws) -> set[str]:
    out: set[str] = set()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                out.add(cell.value.strip())
    return out


def test_artifact_is_workbook_with_keys() -> None:
    # The harness saves the artifact without an .xlsx extension; load from bytes so openpyxl judges it
    # on content. A renamed CSV is not a valid zip container and raises here.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(output().read_bytes()), data_only=True)
    except Exception as exc:  # noqa: BLE001 -- any load failure means it is not a real workbook
        raise AssertionError(f"artifact is not a readable .xlsx workbook: {exc!r}") from exc

    assert len(wb.worksheets) >= 1, "workbook has no worksheets"

    expected = _expected()
    all_text: set[str] = set()
    all_numbers: list[float] = []
    for ws in wb.worksheets:
        all_text |= _text_cells(ws)
        all_numbers += _numeric_cell_values(ws)

    # The workbook must carry the two breakdowns the user asked for: every trade name and every level
    # name appears as a text cell. These are labels, not computed values, so they're immune to whether
    # the model wrote totals as literals or as live =SUM() formulas (openpyxl reads uncached formulas
    # as None under data_only). Correctness of the numbers themselves is graded authoritatively by the
    # result JSON above; here we only confirm a real, populated summary -- not a labels-only skeleton
    # and not a renamed CSV.
    missing_trades = sorted(t for t in expected["totals_by_trade"] if t not in all_text)
    assert not missing_trades, (
        f"trade name(s) {missing_trades} not found as text cells in the workbook; text cells: {sorted(all_text)}"
    )
    missing_levels = sorted(lvl for lvl in expected["totals_by_level"] if lvl not in all_text)
    assert not missing_levels, (
        f"level name(s) {missing_levels} not found as text cells in the workbook; text cells: {sorted(all_text)}"
    )
    # A populated summary, not a labels-only skeleton with a single dummy number: at least one numeric
    # cell per trade and per level (the breakdowns asked for). Counting cells -- not matching specific
    # totals -- keeps this immune to whether the model wrote literals or =SUM() formulas.
    min_numbers = len(expected["totals_by_trade"]) + len(expected["totals_by_level"])
    assert len(all_numbers) >= min_numbers, (
        f"workbook has only {len(all_numbers)} numeric cells; expected at least {min_numbers} for a "
        f"populated by-trade and by-level takeoff summary"
    )
