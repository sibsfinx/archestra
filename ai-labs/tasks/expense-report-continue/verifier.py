"""Grade the continued March expense report.

The march transactions live only in the report the agent saved in the first chat; the second chat runs
in a fresh conversation with an empty sandbox and no attachments. So a final workbook that reproduces
those transactions can only have come from finding and reopening that saved report -- there is nowhere
else the data exists. That makes the line-item fidelity check (below) the real gate on the
cross-conversation "pull up my report" behavior, not a proxy for it.

Reads BENCH_OUTPUT (the exported .xlsx), BENCH_RESULT (the submitted JSON), and the never-staged
BENCH_FIXTURES/expected/report.json (the ground-truth line items and totals). Rows are matched on the
category+amount pair (unique across the eight transactions), so the check does not depend on how the
agent orders columns or formats dates.
"""

import io
import re

import openpyxl

from bench_verifier import output, read_fixture_json, result

MONEY_TOL = 0.01

TRUTHY = {"yes", "y", "true", "t", "reimbursable", "1", "1.0", "✓", "x"}
FALSY = {"no", "n", "false", "f", "not reimbursable", "non-reimbursable", "0", "0.0", "", "-", "none"}


# === helpers ===

def _expected() -> dict:
    return read_fixture_json("expected", "report.json")


def _cents(value) -> int | None:
    """Amount -> integer cents, tolerant of currency symbols and thousands separators."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return round(float(value) * 100)
    if isinstance(value, str):
        cleaned = re.sub(r"[^0-9.\-]", "", value.replace(",", ""))
        if cleaned in ("", "-", ".", "-."):
            return None
        try:
            return round(float(cleaned) * 100)
        except ValueError:
            return None
    return None


def _norm(value) -> str:
    return re.sub(r"\s+", " ", str(value if value is not None else "")).strip().lower()


def _norm_vendor(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower().replace("&", "and")).strip()


def _parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    token = _norm(value)
    if token in TRUTHY:
        return True
    if token in FALSY:
        return False
    raise AssertionError(f"reimbursable cell is not a clear yes/no: {value!r}")


def _workbook():
    return openpyxl.load_workbook(io.BytesIO(output().read_bytes()), data_only=True)


def _header_map(ws) -> dict[str, int] | None:
    """Map logical column -> index from a sheet's first row, or None if it isn't the line-item sheet."""
    header = [_norm(c.value) for c in ws[1]] if ws.max_row else []
    cols: dict[str, int] = {}
    for idx, name in enumerate(header):
        if not name:
            continue
        if "date" in name and "date" not in cols:
            cols["date"] = idx
        elif ("vendor" in name or "merchant" in name) and "vendor" not in cols:
            cols["vendor"] = idx
        elif ("description" in name or "detail" in name or "memo" in name) and "description" not in cols:
            cols["description"] = idx
        elif "categ" in name and "category" not in cols:
            cols["category"] = idx
        elif ("amount" in name or "cost" in name or "total" in name) and "amount" not in cols:
            cols["amount"] = idx
        elif "reimburs" in name and "reimbursable" not in cols:
            cols["reimbursable"] = idx
    # The line-item sheet is the one carrying per-transaction detail: a date, an amount, and a
    # description column. The category-totals sheet (category + total only) won't match.
    if {"date", "amount", "description"} <= cols.keys():
        return cols
    return None


def _line_item_sheet(wb):
    for ws in wb.worksheets:
        cols = _header_map(ws)
        if cols is not None:
            return ws, cols
    raise AssertionError("no worksheet with date/description/amount headers (the line-item report)")


def _data_rows(ws, cols: dict[str, int]) -> list[dict]:
    rows = []
    for r in range(2, ws.max_row + 1):
        values = [c.value for c in ws[r]]
        amount = _cents(values[cols["amount"]]) if cols["amount"] < len(values) else None
        category = values[cols["category"]] if cols.get("category", 99) < len(values) else None
        if amount is None or not _norm(category):
            continue  # blank row or a subtotal/label row without a real category+amount
        rows.append(
            {
                "category": _norm(category),
                "cents": amount,
                "vendor": values[cols["vendor"]] if cols.get("vendor", 99) < len(values) else None,
                "reimbursable": values[cols["reimbursable"]] if cols.get("reimbursable", 99) < len(values) else None,
                "row": r,
            }
        )
    return rows


# === checks ===

def test_line_items_recovered_and_extended() -> None:
    """Every March transaction is present (proving the saved report was found and reopened), with the
    correct vendor and a correct reimbursable flag, and nothing extra."""
    exp = _expected()["line_items"]
    ws, cols = _line_item_sheet(_workbook())
    assert "reimbursable" in cols, "the line-item sheet has no reimbursable column"

    rows = _data_rows(ws, cols)
    by_key: dict[tuple[str, int], dict] = {}
    for row in rows:
        key = (row["category"], row["cents"])
        assert key not in by_key, f"duplicate transaction row for {key} at row {row['row']}"
        by_key[key] = row

    expected_keys = {(_norm(i["category"]), round(i["amount"] * 100)) for i in exp}
    got_keys = set(by_key)
    assert got_keys == expected_keys, (
        f"line items {sorted(got_keys)} != expected {sorted(expected_keys)}; "
        f"missing {sorted(expected_keys - got_keys)}; extra {sorted(got_keys - expected_keys)}"
    )

    for item in exp:
        row = by_key[(_norm(item["category"]), round(item["amount"] * 100))]
        assert _norm_vendor(row["vendor"]) == _norm_vendor(item["vendor"]), (
            f"row {row['row']} vendor {row['vendor']!r} != expected {item['vendor']!r}"
        )
        assert _parse_bool(row["reimbursable"]) == item["reimbursable"], (
            f"row {row['row']} ({item['category']} {item['amount']}) reimbursable "
            f"{row['reimbursable']!r} != expected {item['reimbursable']}"
        )


def test_category_totals_sheet_present() -> None:
    """A second sheet lists per-category totals (numeric correctness is graded via submit_result)."""
    wb = _workbook()
    line_ws, _ = _line_item_sheet(wb)
    others = [ws for ws in wb.worksheets if ws.title != line_ws.title]
    assert others, "expected a second worksheet with per-category totals"

    expected_categories = {_norm(c) for c in _expected()["category_totals"]}
    for ws in others:
        seen = {_norm(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None}
        if expected_categories <= seen:
            return
    raise AssertionError(
        f"no second sheet lists all categories {sorted(expected_categories)}"
    )


def test_submitted_totals() -> None:
    exp = _expected()
    submitted = result()["category_totals"]
    got = {_norm(item["category"]): float(item["total"]) for item in submitted}
    want = {_norm(k): v for k, v in exp["category_totals"].items()}
    assert set(got) == set(want), f"submitted categories {sorted(got)} != expected {sorted(want)}"
    for cat, total in want.items():
        assert abs(got[cat] - total) <= MONEY_TOL, f"category {cat} total {got[cat]} != expected {total}"

    reimbursable = float(result()["reimbursable_total"])
    assert abs(reimbursable - exp["reimbursable_total"]) <= MONEY_TOL, (
        f"reimbursable_total {reimbursable} != expected {exp['reimbursable_total']}"
    )
